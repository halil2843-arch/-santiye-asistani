"""Excel (xlsx) sablon doldurucu.

Kullanim:
    cikti = fill_xlsx(
        sablon_yolu="./uploads/sablonlar/gunluk.xlsx",
        alan_esleme={"B3": "tarih", "D5": "hava_sabah", "F8": "jcb_saat"},
        sonuc=extraction_sonucu,
        cikti_yolu="./outputs/site1/rapor_20260506_a3b4c5d6.xlsx",
    )
"""

from __future__ import annotations

import logging
import os
import re
import shutil
from copy import copy as _copy
from datetime import date
from typing import Any

import openpyxl

from .schemas import ExtractionSonucu

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Alan cozumleme yardimcilari
# ---------------------------------------------------------------------------

_INDEKS_PATTERN = re.compile(
    r"^(personel|makine|is|malzeme)_(\d+)_(.+)$"
)
_IS_FIRMA_PATTERN = re.compile(r"^(.+)_is_(\d+)_firma$")
_BASIT_REF_PATTERN = re.compile(r"^=([A-Z]+\d+)$")  # =D30 gibi tek hücre referansı
_TARIH_SEKME_PATTERN = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})$")  # DD.MM.YYYY sekme adı

# Metin içerikli alanlar — boş kalırsa 0 yerine "" yazılır
_METIN_ALAN_PREFIXLER = {"hava_", "fotograf_", "tarih"}
_METIN_ALAN_SABIT = {"hava_durumu", "hava_genel", "hava_sabah", "hava_ogleden_sonra",
                     "hava_sicaklik", "fotograf_analizi", "tarih", "santiye_adi"}

# Altyüklenici ekip → iş kategorisi eşlemesi
# Puantaj çizelgesi: {ekip_adi}_{rol} → ilgili kategorilerden rol sayısı toplanır
_EKIP_KATEGORI_MAP: dict[str, list[str]] = {
    "kaba_insaat": ["betonarme", "mobilizasyon"],
    "hafriyat":    ["kazi"],
    "izolasyon":   ["altyapi"],
    "cephe_cati":  ["ince_yapi"],
    "mekanik":     ["mekanik"],
    "elektrik":    ["elektrik"],
    "isg":         ["isg"],
}

# Rol anahtarı → YapilanIs alanı
_ROL_ALAN_MAP: dict[str, str] = {
    "usta":      "usta_sayisi",
    "kalipci":   "usta_sayisi",
    "demirci":   "usta_sayisi",
    "beden":     "duz_isci_sayisi",
    "duz_isci":  "duz_isci_sayisi",
    "formen":    "formen_sayisi",
    "kalfa":     "formen_sayisi",
}

# Rol adı → iş açıklamasında bulunması gereken anahtar kelimeler (normalize edilmiş)
_ROL_ACIKLAMA_FILTRE: dict[str, list[str]] = {
    "kalipci": ["kalip"],
    "demirci": ["demir"],
}
# Rol adı → iş açıklamasında OLMAMASI gereken anahtar kelimeler
_ROL_ACIKLAMA_DISLA: dict[str, list[str]] = {
    "usta": ["kalip", "demir"],  # usta = kalıp/demir dışındaki uzman işler
}

_EKIP_ROL_PATTERN = re.compile(
    r"^(" + "|".join(_EKIP_KATEGORI_MAP.keys()) + r")_(.+)$"
)


def _normalize(s: str) -> str:
    """Karşılaştırma için string'i normalize et: küçük harf, Türkçe harf → ASCII, alt çizgi → boşluk."""
    s = s.lower().replace("_", " ").replace("-", " ").strip()
    return (s.replace("ı", "i").replace("ğ", "g").replace("ş", "s")
             .replace("ü", "u").replace("ö", "o").replace("ç", "c"))


def _bos_deger(alan_adi: str) -> Any:
    """Veri yoksa hücreye yazılacak varsayılan değer.

    - Sayısal alanlar (personel sayısı, makine saati) → 0
    - Metin alanlar (açıklama, hava durumu, tarih) → "" (placeholder temizlenir)
    """
    if alan_adi in _METIN_ALAN_SABIT:
        return ""
    if re.match(r"^.+_is_\d+", alan_adi):  # _is_1, _is_2, _is_1_firma vb.
        return ""
    return 0  # personel sayıları, makine saatleri


def _alan_degerini_coz(alan_adi: str, sonuc: ExtractionSonucu) -> Any:
    """Alan adını ExtractionSonucu'ndan karşılık gelen değere dönüştürür.

    Öncelik sırası:
      1. Sabit tekil alanlar (tarih, hava_* vb.)
      2. N-indexli liste alanları  (personel_0_sayi, makine_1_saat vb.)
      3. Semantik makine arama    (jcb_saat, fore_kazik_saat vb.)
      4. Semantik iş arama        (mobilizasyon_is_1, kazi_is_2 vb.)
      5. Semantik iş firma/çalışan (mobilizasyon_is_1_firma vb.)
      6. Semantik personel arama  (formen, proje_muduru_saha vb.)
      7. Ekip puantaj arama       (kaba_insaat_usta, hafriyat_beden vb.)
    """
    # --- 1. Sabit tekil alanlar ---
    if alan_adi == "santiye_adi":
        return sonuc.santiye_adi
    if alan_adi == "tarih":
        return sonuc.tarih.strftime("%d.%m.%Y")
    if alan_adi in ("hava_durumu", "hava_genel"):
        return sonuc.hava_durumu.genel_aciklama
    if alan_adi == "hava_sabah":
        return sonuc.hava_durumu.sabah
    if alan_adi == "hava_ogleden_sonra":
        return sonuc.hava_durumu.ogleden_sonra
    if alan_adi == "hava_sicaklik":
        return sonuc.hava_durumu.sicaklik_derece
    if alan_adi == "fotograf_analizi":
        return sonuc.fotograf_analizi
    if alan_adi == "toplam_personel":
        return sum(p.sayi for p in sonuc.personel)

    # --- 2. N-indexli liste alanları (personel_0_sayi, makine_1_saat ...) ---
    m = _INDEKS_PATTERN.match(alan_adi)
    if m:
        grup, idx_str, alt_alan = m.group(1), m.group(2), m.group(3)
        idx = int(idx_str)

        if grup == "personel":
            if idx >= len(sonuc.personel):
                return None
            kayit = sonuc.personel[idx]
            if alt_alan == "ekip":
                return kayit.ekip_adi
            if alt_alan == "meslek":
                return kayit.meslek
            if alt_alan == "sayi":
                return kayit.sayi

        elif grup == "makine":
            if idx >= len(sonuc.makineler):
                return None
            kayit = sonuc.makineler[idx]
            if alt_alan == "tipi":
                return kayit.makine_tipi
            if alt_alan == "sayi":
                return kayit.sayi
            if alt_alan == "saat":
                return kayit.calisma_saati

        elif grup == "is":
            if idx >= len(sonuc.yapilan_isler):
                return None
            kayit = sonuc.yapilan_isler[idx]
            if alt_alan == "kategori":
                return kayit.kategori
            if alt_alan == "aciklama":
                return kayit.aciklama

        elif grup == "malzeme":
            if idx >= len(sonuc.malzeme_girisi):
                return None
            kayit = sonuc.malzeme_girisi[idx]
            if alt_alan == "adi":
                return kayit.malzeme_adi
            if alt_alan == "miktar":
                return kayit.miktar
            if alt_alan == "birim":
                return kayit.birim

        return None

    # --- 3. Semantik makine arama: {makine_adi}_saat ---
    # Örnek: jcb_saat → makineler listesinde "jcb" geçen kaydın calisma_saati
    if alan_adi.endswith("_saat"):
        anahtar = _normalize(alan_adi[:-5])  # "_saat" kısmını çıkar
        for makine in sonuc.makineler:
            if anahtar in _normalize(makine.makine_tipi):
                # Saat bilinmiyorsa makine adedini yaz (sahada mevcut sinyali)
                return makine.calisma_saati if makine.calisma_saati is not None else makine.sayi
        return None

    # --- 4. Semantik iş arama: {kategori}_is_{N} (1-tabanlı) ---
    # Örnek: mobilizasyon_is_1 → "mobilizasyon" kategorisindeki 1. iş açıklaması
    is_m = re.match(r"^(.+)_is_(\d+)$", alan_adi)
    if is_m:
        kategori_anahtar = _normalize(is_m.group(1))
        idx = int(is_m.group(2)) - 1  # 1-tabanlı → 0-tabanlı
        eslesme = [
            is_ for is_ in sonuc.yapilan_isler
            if kategori_anahtar in _normalize(is_.kategori)
        ]
        if 0 <= idx < len(eslesme):
            return eslesme[idx].aciklama
        return None

    # --- 5. Semantik iş firma/çalışan: {kategori}_is_{N}_firma ---
    # Örnek: kazi_is_1_firma → "kazi" kategorisindeki 1. işin firma/çalışan bilgisi
    firma_m = _IS_FIRMA_PATTERN.match(alan_adi)
    if firma_m:
        kategori_anahtar = _normalize(firma_m.group(1))
        idx = int(firma_m.group(2)) - 1
        eslesme = [
            is_ for is_ in sonuc.yapilan_isler
            if kategori_anahtar in _normalize(is_.kategori)
        ]
        if 0 <= idx < len(eslesme):
            is_ = eslesme[idx]
            parcalar: list[str] = []
            if is_.ilgili_firma:
                parcalar.append(is_.ilgili_firma)
            if is_.calisan_sayisi is not None:
                parcalar.append(str(is_.calisan_sayisi))
            return " / ".join(parcalar) if parcalar else None
        return None

    # --- 6. Semantik personel arama: rol/meslek adıyla ---
    # Örnek: formen → personel listesinde meslek "formen" olan kaydın sayısı
    rol_anahtar = _normalize(alan_adi)
    for p in sonuc.personel:
        if rol_anahtar in _normalize(p.meslek) or rol_anahtar in _normalize(p.ekip_adi):
            return p.sayi

    # --- 7. Ekip puantaj arama: {ekip}_{rol} ---
    # Örnek: kaba_insaat_usta → betonarme/mobilizasyon işlerindeki usta_sayisi toplamı
    # Örnek: hafriyat_beden → kazi işlerindeki duz_isci_sayisi toplamı
    ekip_m = _EKIP_ROL_PATTERN.match(alan_adi)
    if ekip_m:
        ekip_adi = ekip_m.group(1)
        rol_adi = ekip_m.group(2)
        kategoriler = _EKIP_KATEGORI_MAP.get(ekip_adi, [])
        # Rol adını YapilanIs alanına çevir
        alan_adi_rol = None
        for anahtar, deger in _ROL_ALAN_MAP.items():
            if anahtar in rol_adi:
                alan_adi_rol = deger
                break
        if alan_adi_rol and kategoriler:
            aciklama_filtre = _ROL_ACIKLAMA_FILTRE.get(rol_adi)
            aciklama_disla = _ROL_ACIKLAMA_DISLA.get(rol_adi)
            # Aynı alt-kategorideki birden fazla iş aynı işçileri sayabilir,
            # bu yüzden her alt-kategori için max alıp farklı kategorileri topluyoruz.
            maks_per_kat: dict[str, int] = {}
            for is_ in sonuc.yapilan_isler:
                is_kat = _normalize(is_.kategori)
                eslen_kat = next((k for k in kategoriler if k in is_kat), None)
                if not eslen_kat:
                    continue
                aciklama_norm = _normalize(is_.aciklama or "")
                if aciklama_filtre and not any(f in aciklama_norm for f in aciklama_filtre):
                    continue
                if aciklama_disla and any(f in aciklama_norm for f in aciklama_disla):
                    continue
                deger = getattr(is_, alan_adi_rol, None)
                if deger is not None:
                    maks_per_kat[eslen_kat] = max(maks_per_kat.get(eslen_kat, 0), deger)
            toplam = sum(maks_per_kat.values())
            return toplam if toplam > 0 else None

    logger.debug("Tanımsız alan adı, sessizce atlandı: '%s'", alan_adi)
    return None


# ---------------------------------------------------------------------------
# Ana fonksiyon
# ---------------------------------------------------------------------------


def _kumulatif_adresler_bul(sablon_yolu: str) -> dict[str, str]:
    """Şablondaki =XN formüllü hücreleri {kümülatif_adres: kaynak_adres} olarak döndürür.

    Sadece tam hücre referansı olan formüller (=D30, =H8 vb.) eşleşir;
    =SUM(…) veya karmaşık formüller dahil değildir.
    """
    wb = openpyxl.load_workbook(sablon_yolu)
    ws = wb.active
    kumul: dict[str, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                m = _BASIT_REF_PATTERN.match(cell.value)
                if m:
                    kumul[cell.coordinate] = m.group(1)
    return kumul


def _sekme_tarih(sekme_adi: str) -> date:
    """Sekme adını (DD.MM.YYYY) date nesnesine çevirir; geçersizse date.min döner."""
    m = _TARIH_SEKME_PATTERN.match(sekme_adi)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return date.min


def _kumulatif_guncelle(wb: openpyxl.Workbook, kumul_map: dict[str, str]) -> None:
    """Tüm sekmelerdeki kümülatif hücreleri tarihe göre sıralı kümültatif toplamla günceller.

    Her DD.MM.YYYY sekmesi için kümülatif değer = o tarihe kadarki tüm sekmelerin
    aynı kaynak hücresindeki günlük değerlerinin toplamı.
    """
    if not kumul_map:
        return

    sirali = sorted(wb.sheetnames, key=_sekme_tarih)

    for i, sekme_adi in enumerate(sirali):
        ws = wb[sekme_adi]
        for kumul_adres, kaynak_adres in kumul_map.items():
            toplam = 0
            for j in range(i + 1):  # 0..i (bu güne kadar tüm günler dahil)
                val = wb[sirali[j]][kaynak_adres].value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    toplam += val
            ws[kumul_adres] = toplam


def _sablon_sayfasini_kopyala(ws_sablon, wb_hedef, sekme_adi: str):
    """Şablon sayfasını hedef workbook'a yeni sekme olarak kopyalar (stil + boyut + birleştirme dahil)."""
    ws_yeni = wb_hedef.create_sheet(title=sekme_adi)

    for row in ws_sablon.iter_rows():
        for cell in row:
            new_cell = ws_yeni.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.border = _copy(cell.border)
                new_cell.fill = _copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = _copy(cell.protection)
                new_cell.alignment = _copy(cell.alignment)

    for col_harf, col_dim in ws_sablon.column_dimensions.items():
        ws_yeni.column_dimensions[col_harf].width = col_dim.width
        ws_yeni.column_dimensions[col_harf].hidden = col_dim.hidden

    for row_idx, row_dim in ws_sablon.row_dimensions.items():
        ws_yeni.row_dimensions[row_idx].height = row_dim.height
        ws_yeni.row_dimensions[row_idx].hidden = row_dim.hidden

    for birlesim in ws_sablon.merged_cells.ranges:
        ws_yeni.merge_cells(str(birlesim))

    return ws_yeni


def fill_xlsx(
    sablon_yolu: str,
    alan_esleme: dict[str, str],  # {"B3": "tarih", "D5": "hava_sabah", ...}
    sonuc: ExtractionSonucu,
    cikti_yolu: str,
    llm_mapping: dict | None = None,
) -> str:
    """Excel şablonunu ExtractionSonucu verileriyle doldurur.

    Aylık dosya mantığı:
      - cikti_yolu dosyası yoksa: şablonu kopyala, ilk sekmeyi tarihle adlandır.
      - cikti_yolu dosyası varsa: şablondan yeni sekme kopyala, aylık dosyaya ekle.

    Args:
        sablon_yolu:  Kaynak xlsx şablonunun disk yolu.
        alan_esleme:  Hücre adresi → alan adı eşleme sözlüğü.
        sonuc:        Aşama 1 extraction sonucu.
        cikti_yolu:   Aylık çıktı dosyasının tam yolu (rapor_YYYYMM.xlsx).
        llm_mapping:  Aşama 2 LLM eşleme sonucu. None ise sadece heuristic.

    Returns:
        cikti_yolu
    """
    tarih_str = sonuc.tarih.strftime("%d.%m.%Y")

    if not os.path.exists(cikti_yolu):
        # Ayın ilk raporu — şablonu kopyala, aktif sayfayı tarihle adlandır
        shutil.copy2(sablon_yolu, cikti_yolu)
        logger.info("Aylık dosya oluşturuldu: %s", cikti_yolu)
        wb = openpyxl.load_workbook(cikti_yolu)
        ws = wb.active
        ws.title = tarih_str

        # Boş ekstra sayfaları sil (şablon bazen birden fazla sayfa içerebilir)
        for sayfa_adi in list(wb.sheetnames):
            if sayfa_adi == tarih_str:
                continue
            bos = wb[sayfa_adi]
            if not any(c.value is not None for r in bos.iter_rows() for c in r):
                del wb[sayfa_adi]
                logger.info("Boş sayfa silindi: '%s'", sayfa_adi)
    else:
        # Aylık dosya mevcut — şablondan yeni sekme ekle
        wb = openpyxl.load_workbook(cikti_yolu)

        if tarih_str in wb.sheetnames:
            del wb[tarih_str]
            logger.info("Mevcut sekme yeniden üretiliyor: '%s'", tarih_str)

        wb_sablon = openpyxl.load_workbook(sablon_yolu)
        ws = _sablon_sayfasini_kopyala(wb_sablon.active, wb, tarih_str)
        logger.info("Yeni sekme eklendi: '%s' → %s", tarih_str, cikti_yolu)

    # Sekmeyi doldur
    yazilan = 0
    atlanan = 0

    for hucre_adresi, alan_adi in alan_esleme.items():
        if llm_mapping and alan_adi in llm_mapping and llm_mapping[alan_adi] is not None:
            deger = llm_mapping[alan_adi]
            logger.debug("LLM mapping — hucre=%s alan=%s deger=%s", hucre_adresi, alan_adi, deger)
        else:
            deger = _alan_degerini_coz(alan_adi, sonuc)
        if deger is None:
            deger = _bos_deger(alan_adi)
        try:
            ws[hucre_adresi] = deger
            yazilan += 1
            logger.debug("Yazildi — hucre=%s alan=%s deger=%s", hucre_adresi, alan_adi, deger)
        except (KeyError, ValueError) as exc:
            logger.warning("Hucre yazma hatasi hucre=%s: %s", hucre_adresi, exc)
            atlanan += 1

    # Kümülatif sütunları tüm sekmeler üzerinden yeniden hesapla
    kumul_map = _kumulatif_adresler_bul(sablon_yolu)
    _kumulatif_guncelle(wb, kumul_map)
    logger.info(
        "Kümülatif güncellendi — %d hücre × %d sekme",
        len(kumul_map), len(wb.sheetnames),
    )

    wb.save(cikti_yolu)
    logger.info(
        "Excel raporu kaydedildi: %s  (sekme=%s, yazilan=%d, atlanan=%d)",
        cikti_yolu, tarih_str, yazilan, atlanan,
    )
    return cikti_yolu
