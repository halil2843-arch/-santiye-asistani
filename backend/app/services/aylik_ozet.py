"""
Aylık özet Excel raporu.
Her ay 1'inde APScheduler çalıştıracak (Faz 3'te).
"""
from __future__ import annotations

import os
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def aylik_ozet_olustur(
    santiye_adi: str,
    ay: int,
    yil: int,
    gunluk_veriler: list[dict],  # [{"tarih": "2026-05-01", "personel": 24, "makine": 5, ...}]
    cikti_dizin: str = "./outputs",
) -> str:
    """
    Aylık özet Excel raporu oluşturur.

    Args:
        santiye_adi:    Şantiye ismi (dosya adı ve başlıkta kullanılır).
        ay:             Ay numarası (1-12).
        yil:            Yıl (örn. 2026).
        gunluk_veriler: Her günün rapor özetini içeren liste.
                        Beklenen anahtarlar:
                          - tarih    : str  "YYYY-MM-DD"
                          - personel : int  Sahada toplam kişi
                          - makine   : int  Aktif makine adedi
                          - is_kalemi: str  Günün ana iş kalemi
                          - hava     : str  Hava durumu özeti
                          - durum    : str  "taslak" | "onaylandi" | "hata"
                          - notlar   : str  Ek notlar
        cikti_dizin:    Dosyanın kaydedileceği dizin yolu.

    Returns:
        Üretilen dosyanın tam yolu.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{yil}-{ay:02d} Özet"

    # -----------------------------------------------------------------------
    # Başlık satırı (A1:G1)
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = f"AYLIK ÖZET RAPORU — {santiye_adi} — {yil}/{ay:02d}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E2636")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # -----------------------------------------------------------------------
    # Tablo başlıkları (satır 3)
    # -----------------------------------------------------------------------
    basliklar = [
        "Tarih",
        "Personel",
        "Makine",
        "Ana İş Kalemi",
        "Hava",
        "Rapor Durumu",
        "Notlar",
    ]
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=3, column=col, value=baslik)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 18

    # -----------------------------------------------------------------------
    # Veri satırları
    # -----------------------------------------------------------------------
    toplam_personel = 0
    toplam_makine = 0

    for i, gun in enumerate(gunluk_veriler, 1):
        row = 3 + i
        personel = gun.get("personel", 0) or 0
        makine = gun.get("makine", 0) or 0
        toplam_personel += personel
        toplam_makine += makine

        ws.cell(row=row, column=1, value=gun.get("tarih", ""))
        ws.cell(row=row, column=2, value=personel)
        ws.cell(row=row, column=3, value=makine)
        ws.cell(row=row, column=4, value=gun.get("is_kalemi", ""))
        ws.cell(row=row, column=5, value=gun.get("hava", ""))
        ws.cell(row=row, column=6, value=gun.get("durum", ""))
        ws.cell(row=row, column=7, value=gun.get("notlar", ""))

        # Renk kodlaması: onaylı → açık yeşil, hata → açık kırmızı
        durum = gun.get("durum", "")
        if durum == "onaylandi":
            satir_renk = "E8F5E9"
        elif durum == "hata":
            satir_renk = "FFEBEE"
        elif i % 2 == 0:
            satir_renk = "F8FAFC"
        else:
            satir_renk = "FFFFFF"

        for c in range(1, 8):
            hucre = ws.cell(row=row, column=c)
            hucre.fill = PatternFill("solid", fgColor=satir_renk)
            hucre.alignment = Alignment(vertical="top", wrap_text=(c in (4, 7)))

    # -----------------------------------------------------------------------
    # Özet satır
    # -----------------------------------------------------------------------
    son_veri_satiri = 3 + len(gunluk_veriler)
    ozet_row = son_veri_satiri + 2

    ws.cell(row=ozet_row, column=1, value="TOPLAM / ORTALAMA").font = Font(bold=True)
    ws.cell(row=ozet_row, column=1).fill = PatternFill("solid", fgColor="E2E8F0")

    toplam_p_cell = ws.cell(row=ozet_row, column=2)
    toplam_p_cell.value = toplam_personel
    toplam_p_cell.font = Font(bold=True, color="F59E0B")
    toplam_p_cell.fill = PatternFill("solid", fgColor="E2E8F0")

    toplam_m_cell = ws.cell(row=ozet_row, column=3)
    toplam_m_cell.value = toplam_makine
    toplam_m_cell.font = Font(bold=True)
    toplam_m_cell.fill = PatternFill("solid", fgColor="E2E8F0")

    gun_sayisi = len(gunluk_veriler)
    if gun_sayisi > 0:
        ort_personel_cell = ws.cell(
            row=ozet_row + 1, column=2,
            value=round(toplam_personel / gun_sayisi, 1),
        )
        ort_personel_cell.font = Font(italic=True, color="64748B")
        ws.cell(row=ozet_row + 1, column=1, value="Günlük Ort. Personel").font = Font(
            italic=True, color="64748B"
        )

    # -----------------------------------------------------------------------
    # Kolon genişlikleri
    # -----------------------------------------------------------------------
    genislikler = [12, 10, 10, 35, 20, 16, 30]
    for col, gen in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(col)].width = gen

    # -----------------------------------------------------------------------
    # Kaydet
    # -----------------------------------------------------------------------
    os.makedirs(cikti_dizin, exist_ok=True)
    guvenli_ad = santiye_adi.replace(" ", "_").replace("/", "-")
    dosya = f"aylik_ozet_{guvenli_ad}_{yil}_{ay:02d}.xlsx"
    yol = os.path.join(cikti_dizin, dosya)
    wb.save(yol)
    return yol
