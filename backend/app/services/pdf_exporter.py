"""
Excel raporunu PDF'e dönüştürür.
WeasyPrint veya pdfkit yoksa HTML olarak kaydeder (tarayıcıdan yazdırılabilir).
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Renk yardımcısı
# ---------------------------------------------------------------------------

def _hucre_arkaplan(cell) -> str | None:
    """Hücrenin dolgu rengini hex string olarak döndürür, yoksa None."""
    try:
        fill = cell.fill
        if fill and fill.fill_type == "solid":
            argb = fill.fgColor.rgb  # "FF1E2636" formatı
            if argb and len(argb) == 8 and argb != "00000000":
                return f"#{argb[2:]}"  # alpha'yı at
    except Exception:
        pass
    return None


def _hucre_metin_rengi(cell) -> str | None:
    """Hücrenin yazı rengini hex string olarak döndürür, yoksa None."""
    try:
        font = cell.font
        if font and font.color:
            argb = font.color.rgb
            if argb and len(argb) == 8 and argb != "FF000000":
                return f"#{argb[2:]}"
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Excel → HTML çevrimi
# ---------------------------------------------------------------------------

def excel_to_html(xlsx_yolu: str) -> str:
    """Excel dosyasını stil korumalı HTML tablosuna çevirir."""
    wb = openpyxl.load_workbook(xlsx_yolu, data_only=True)

    html_parts = [
        "<!DOCTYPE html><html><head>",
        '<meta charset="UTF-8">',
        "<style>",
        "body { font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }",
        "table { border-collapse: collapse; width: 100%; margin-bottom: 30px; }",
        "th { background: #1E2636; color: white; padding: 6px 8px; text-align: left;"
        " border: 1px solid #334155; }",
        "td { border: 1px solid #ddd; padding: 4px 8px; vertical-align: top; }",
        "tr:nth-child(even) td { background: #f8f9fa; }",
        "h2 { color: #1E2636; border-bottom: 2px solid #F59E0B; padding-bottom: 4px;"
        " margin-top: 30px; }",
        ".bold { font-weight: bold; }",
        ".center { text-align: center; }",
        ".right { text-align: right; }",
        "@media print { .no-print { display: none; } body { margin: 0; } }",
        "</style></head><body>",
        '<div class="no-print" style="margin-bottom:16px;">',
        '<button onclick="window.print()" style="padding:8px 16px;background:#1E2636;'
        "color:white;border:none;cursor:pointer;font-size:13px;\">",
        "🖨️ Yazdır / PDF Olarak Kaydet</button></div>",
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_parts.append(f"<h2>{sheet_name}</h2><table>")

        max_satir = ws.max_row or 0
        max_sutun = ws.max_column or 0

        for r_idx, row in enumerate(
            ws.iter_rows(max_row=max_satir, max_col=max_sutun, values_only=False), 1
        ):
            if all(c.value is None for c in row):
                continue

            # İlk satırı <th> olarak işle
            tag = "th" if r_idx == 1 else "td"
            html_parts.append("<tr>")
            for cell in row:
                val = cell.value if cell.value is not None else ""
                # Stil
                stil_parcalar: list[str] = []
                bg = _hucre_arkaplan(cell)
                if bg:
                    stil_parcalar.append(f"background:{bg}")
                fg = _hucre_metin_rengi(cell)
                if fg:
                    stil_parcalar.append(f"color:{fg}")
                stil_attr = f' style="{";".join(stil_parcalar)}"' if stil_parcalar else ""

                # CSS sınıfları
                css_siniflar: list[str] = []
                try:
                    if cell.font and cell.font.bold:
                        css_siniflar.append("bold")
                except Exception:
                    pass
                try:
                    if cell.alignment:
                        h = cell.alignment.horizontal or "left"
                        if h in ("center", "right"):
                            css_siniflar.append(h)
                except Exception:
                    pass
                sinif_attr = f' class="{" ".join(css_siniflar)}"' if css_siniflar else ""

                html_parts.append(f"<{tag}{sinif_attr}{stil_attr}>{val}</{tag}>")
            html_parts.append("</tr>")

        html_parts.append("</table>")

    html_parts.append("</body></html>")
    return "\n".join(html_parts)


# ---------------------------------------------------------------------------
# Kayıt fonksiyonları
# ---------------------------------------------------------------------------

def rapor_html_kaydet(xlsx_yolu: str, cikti_dizin: str | None = None) -> str:
    """Excel'i HTML olarak kaydeder, yolu döndürür."""
    html = excel_to_html(xlsx_yolu)
    xlsx_path = Path(xlsx_yolu)
    cikti = Path(cikti_dizin) if cikti_dizin else xlsx_path.parent
    html_yolu = str(cikti / xlsx_path.stem) + ".html"
    with open(html_yolu, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info("HTML raporu kaydedildi: %s", html_yolu)
    return html_yolu


def rapor_pdf_kaydet(xlsx_yolu: str, cikti_dizin: str | None = None) -> str:
    """Excel'i PDF'e dönüştürür.

    Önce WeasyPrint, sonra pdfkit dener; ikisi de yoksa HTML kaydeder ve
    dönen yolda .html uzantısı bulunur (tarayıcıdan yazdırılabilir).
    """
    html = excel_to_html(xlsx_yolu)
    xlsx_path = Path(xlsx_yolu)
    cikti = Path(cikti_dizin) if cikti_dizin else xlsx_path.parent

    # --- WeasyPrint denemesi ---
    try:
        import weasyprint  # noqa: PLC0415
        pdf_yolu = str(cikti / xlsx_path.stem) + ".pdf"
        weasyprint.HTML(string=html).write_pdf(pdf_yolu)
        logger.info("WeasyPrint PDF üretildi: %s", pdf_yolu)
        return pdf_yolu
    except ImportError:
        logger.debug("WeasyPrint mevcut değil, pdfkit deneniyor.")
    except Exception as exc:
        logger.warning("WeasyPrint hatası: %s, pdfkit deneniyor.", exc)

    # --- pdfkit denemesi ---
    try:
        import pdfkit  # noqa: PLC0415
        pdf_yolu = str(cikti / xlsx_path.stem) + ".pdf"
        pdfkit.from_string(html, pdf_yolu)
        logger.info("pdfkit PDF üretildi: %s", pdf_yolu)
        return pdf_yolu
    except ImportError:
        logger.debug("pdfkit mevcut değil, HTML fallback aktif.")
    except Exception as exc:
        logger.warning("pdfkit hatası: %s, HTML fallback aktif.", exc)

    # --- HTML fallback ---
    logger.info("PDF kütüphanesi bulunamadı, HTML olarak kaydediliyor.")
    return rapor_html_kaydet(xlsx_yolu, cikti_dizin)
