"""
ISG (İş Sağlığı ve Güvenliği) Aylık Özet Raporu — Excel oluşturucu.

Çıktı:
  - Başlık: Şantiye adı + ay/yıl
  - Tablo: Tarih | Denetim Tipi | Bulgular | Düzeltici Faaliyet | Sorumlu | Durum
  - Renk kodlu durum sütunu: Tamamlandı=yeşil, Devam Ediyor=sarı, Gecikmiş=kırmızı
  - Alt kısım: İstatistik özeti
"""
from __future__ import annotations

import os
from typing import Literal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Durum → renk eşlemesi
# ---------------------------------------------------------------------------

DurumTipi = Literal["Tamamlandı", "Devam Ediyor", "Gecikmiş"]

_DURUM_RENKLER: dict[str, str] = {
    "Tamamlandı":   "D1FAE5",   # yeşil (açık)
    "Devam Ediyor": "FEF9C3",   # sarı (açık)
    "Gecikmiş":     "FEE2E2",   # kırmızı (açık)
}
_DURUM_FONT_RENKLER: dict[str, str] = {
    "Tamamlandı":   "065F46",   # koyu yeşil
    "Devam Ediyor": "92400E",   # koyu amber
    "Gecikmiş":     "991B1B",   # koyu kırmızı
}

# ---------------------------------------------------------------------------
# Stil sabitleri
# ---------------------------------------------------------------------------

_RENK_BASLIK    = "1E2636"
_RENK_TABLO_BAŞ = "334155"
_RENK_ALT_BG    = "F1F5F9"
_RENK_ZEBRA     = "F8FAFC"
_RENK_BEYAZ     = "FFFFFF"
_RENK_IST_BG    = "E0F2FE"   # istatistik bölümü

_INCE = Side(style="thin", color="CBD5E1")
_KENAR = Border(left=_INCE, right=_INCE, top=_INCE, bottom=_INCE)


def _doldur(renk: str) -> PatternFill:
    return PatternFill("solid", fgColor=renk)


def _hucre(ws, row: int, col: int, deger=None, *,
           bold: bool = False, italic: bool = False,
           font_color: str = "1E293B", font_size: int = 10,
           fill: PatternFill | None = None,
           align_h: str = "left", align_v: str = "center",
           wrap: bool = False, border: bool = True,
           number_format: str | None = None):
    cell = ws.cell(row=row, column=col, value=deger)
    cell.font = Font(bold=bold, italic=italic, size=font_size,
                     color=font_color, name="Arial")
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v,
                               wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = _KENAR
    if number_format:
        cell.number_format = number_format
    return cell


# ---------------------------------------------------------------------------
# Ana fonksiyon
# ---------------------------------------------------------------------------


def isg_raporu_olustur(
    santiye_adi: str,
    ay_yil: str,           # "Mayıs 2026" veya "2026-05"
    denetimler: list[dict],
    # dict anahtarları:
    #   tarih: str          → "05.05.2026"
    #   denetim_tipi: str   → "Rutin Denetim", "Kaza Araştırması" vb.
    #   bulgular: str       → Bulunan uygunsuzluklar
    #   duz_faaliyeti: str  → Düzeltici faaliyet açıklaması
    #   sorumlu: str        → Sorumlu kişi/birim
    #   durum: DurumTipi    → "Tamamlandı" | "Devam Ediyor" | "Gecikmiş"
    cikti_dizin: str = "./outputs",
) -> str:
    """ISG aylık özet Excel raporu oluşturur.

    Returns:
        Oluşturulan dosyanın tam yolu.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ISG Özet"
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # Satır 1 — Ana başlık
    # ------------------------------------------------------------------
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = f"ISG AYLIK ÖZET RAPORU — {santiye_adi.upper()}"
    cell.font = Font(bold=True, size=14, color=_RENK_BEYAZ, name="Arial")
    cell.fill = _doldur(_RENK_BASLIK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Satır 2 — Dönem
    ws.merge_cells("A2:F2")
    cell2 = ws["A2"]
    cell2.value = f"Dönem: {ay_yil}"
    cell2.font = Font(bold=True, size=11, color=_RENK_BASLIK, name="Arial")
    cell2.fill = _doldur("E2E8F0")
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Satır 3 — boşluk
    ws.row_dimensions[3].height = 8

    # ------------------------------------------------------------------
    # Satır 4 — Tablo başlıkları
    # ------------------------------------------------------------------
    sutun_basliklar = [
        "Tarih", "Denetim Tipi", "Bulgular",
        "Düzeltici Faaliyet", "Sorumlu", "Durum"
    ]
    for col, baslik in enumerate(sutun_basliklar, 1):
        _hucre(ws, 4, col, baslik,
               bold=True, font_color=_RENK_BEYAZ, font_size=10,
               fill=_doldur(_RENK_TABLO_BAŞ), align_h="center")
    ws.row_dimensions[4].height = 22

    # ------------------------------------------------------------------
    # Satır 5+ — Denetim kayıtları
    # ------------------------------------------------------------------
    for i, kayit in enumerate(denetimler, 1):
        row = 4 + i
        durum = kayit.get("durum", "Devam Ediyor")
        durum_renk = _DURUM_RENKLER.get(durum, "FFFFFF")
        durum_font = _DURUM_FONT_RENKLER.get(durum, "000000")
        zebra = _doldur(_RENK_ZEBRA) if i % 2 == 0 else None

        def _satir_hucre(col, deger, **kw):
            if zebra and "fill" not in kw:
                kw["fill"] = zebra
            _hucre(ws, row, col, deger, **kw)

        _satir_hucre(1, kayit.get("tarih", ""), align_h="center")
        _satir_hucre(2, kayit.get("denetim_tipi", ""), align_h="center")
        _satir_hucre(3, kayit.get("bulgular", ""), wrap=True)
        _satir_hucre(4, kayit.get("duz_faaliyeti", ""), wrap=True)
        _satir_hucre(5, kayit.get("sorumlu", ""), align_h="center")

        # Durum — renkli badge etkisi
        _hucre(ws, row, 6, durum,
               bold=True, font_color=durum_font, font_size=10,
               fill=_doldur(durum_renk), align_h="center")

        ws.row_dimensions[row].height = 40

    # ------------------------------------------------------------------
    # İstatistik özeti bölümü
    # ------------------------------------------------------------------
    n = len(denetimler)
    ist_bas_row = n + 6
    toplam_denetim  = n
    tamamlanan      = sum(1 for d in denetimler if d.get("durum") == "Tamamlandı")
    acik_madde      = sum(1 for d in denetimler if d.get("durum") in ("Devam Ediyor", "Gecikmiş"))
    gecikmiş        = sum(1 for d in denetimler if d.get("durum") == "Gecikmiş")

    # Başlık
    ws.merge_cells(f"A{ist_bas_row}:F{ist_bas_row}")
    ist_baslik = ws[f"A{ist_bas_row}"]
    ist_baslik.value = "İSTATİSTİK ÖZETİ"
    ist_baslik.font = Font(bold=True, size=11, color=_RENK_BASLIK, name="Arial")
    ist_baslik.fill = _doldur(_RENK_IST_BG)
    ist_baslik.alignment = Alignment(horizontal="center", vertical="center")
    ist_baslik.border = _KENAR
    ws.row_dimensions[ist_bas_row].height = 22

    istatistikler = [
        ("Toplam Denetim Sayısı",  toplam_denetim),
        ("Tamamlanan Madde",       tamamlanan),
        ("Açık Madde",             acik_madde),
        ("Gecikmiş Madde",         gecikmiş),
    ]

    for j, (etiket, deger) in enumerate(istatistikler):
        ist_row = ist_bas_row + 1 + j

        ws.merge_cells(f"A{ist_row}:C{ist_row}")
        lbl = ws[f"A{ist_row}"]
        lbl.value = etiket
        lbl.font = Font(bold=True, size=10, color=_RENK_BASLIK, name="Arial")
        lbl.fill = _doldur(_RENK_ALT_BG)
        lbl.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)
        lbl.border = _KENAR

        ws.merge_cells(f"D{ist_row}:F{ist_row}")
        val = ws[f"D{ist_row}"]
        val.value = deger
        val.font = Font(bold=True, size=12, color=_RENK_BASLIK, name="Arial")
        val.fill = _doldur(_RENK_ALT_BG)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = _KENAR
        ws.row_dimensions[ist_row].height = 20

    # Tamamlanma oranı (formül)
    oran_row = ist_bas_row + len(istatistikler) + 1
    ws.merge_cells(f"A{oran_row}:C{oran_row}")
    lbl_oran = ws[f"A{oran_row}"]
    lbl_oran.value = "Tamamlanma Oranı"
    lbl_oran.font = Font(bold=True, size=10, color=_RENK_BASLIK, name="Arial")
    lbl_oran.fill = _doldur("D1FAE5")
    lbl_oran.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl_oran.border = _KENAR

    ws.merge_cells(f"D{oran_row}:F{oran_row}")
    val_oran = ws[f"D{oran_row}"]
    # Formül: Tamamlanan / Toplam Denetim (sıfıra bölme korumalı)
    val_oran.value = (
        f"=IF(D{ist_bas_row+1}=0,\"-\","
        f"TEXT(D{ist_bas_row+2}/D{ist_bas_row+1},\"0.0%\"))"
    )
    val_oran.font = Font(bold=True, size=12, color="065F46", name="Arial")
    val_oran.fill = _doldur("D1FAE5")
    val_oran.alignment = Alignment(horizontal="center", vertical="center")
    val_oran.border = _KENAR
    ws.row_dimensions[oran_row].height = 20

    # ------------------------------------------------------------------
    # Kolon genişlikleri
    # ------------------------------------------------------------------
    genislikler = {
        "A": 14,   # Tarih
        "B": 18,   # Denetim Tipi
        "C": 32,   # Bulgular
        "D": 32,   # Düzeltici Faaliyet
        "E": 16,   # Sorumlu
        "F": 16,   # Durum
    }
    for harf, gen in genislikler.items():
        ws.column_dimensions[harf].width = gen

    # ------------------------------------------------------------------
    # Alt bilgi
    # ------------------------------------------------------------------
    footer_row = oran_row + 2
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = (
        f"Şantiye Asistanı ISG Raporu  |  {santiye_adi}  |  {ay_yil}  |  "
        "Bu belge bilgi amaçlıdır, resmi ISG kaydı olarak kabul edilemez."
    )
    footer.font = Font(italic=True, size=8, color="94A3B8", name="Arial")
    footer.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------------------
    # Kaydet
    # ------------------------------------------------------------------
    os.makedirs(cikti_dizin, exist_ok=True)
    ay_yil_dosya = ay_yil.replace(" ", "_").replace("/", "-")
    dosya_adi = f"isg_raporu_{santiye_adi.replace(' ', '_')}_{ay_yil_dosya}.xlsx"
    dosya_yolu = os.path.join(cikti_dizin, dosya_adi)
    wb.save(dosya_yolu)
    return dosya_yolu
