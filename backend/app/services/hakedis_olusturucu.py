"""
Hakediş Excel dosyası oluşturucu.
Aylık raporlardan iş kalemlerini toplayıp hakediş tablosu çıkarır.
"""
from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Yardımcı stil sabitleri
# ---------------------------------------------------------------------------

_RENK_BASLIK     = "1E2636"  # koyu lacivert
_RENK_TABLO_BAŞ  = "334155"  # orta lacivert
_RENK_AMBER      = "F59E0B"  # amber/altın
_RENK_ZEBRA      = "F8FAFC"  # açık mavi-gri (çift satır)
_RENK_BEYAZ      = "FFFFFF"
_RENK_TOPLAM_BG  = "FEF3C7"  # soluk amber arka plan

_INCE_KENAR = Side(style="thin", color="CBD5E1")
_KENAR_TUMU = Border(
    left=_INCE_KENAR, right=_INCE_KENAR,
    top=_INCE_KENAR,  bottom=_INCE_KENAR,
)


def _doldur(renk: str) -> PatternFill:
    return PatternFill("solid", fgColor=renk)


def _yaz(ws, row: int, col: int, deger, **kwargs):
    """Hücreye yaz ve isteğe bağlı stil uygula."""
    cell = ws.cell(row=row, column=col, value=deger)
    if kwargs.get("bold"):
        cell.font = Font(bold=True, color=kwargs.get("font_color", "000000"),
                         size=kwargs.get("font_size", 11), name="Arial")
    else:
        cell.font = Font(color=kwargs.get("font_color", "000000"),
                         size=kwargs.get("font_size", 11), name="Arial")
    if "fill" in kwargs:
        cell.fill = kwargs["fill"]
    if "align" in kwargs:
        cell.alignment = Alignment(horizontal=kwargs["align"], vertical="center",
                                   wrap_text=kwargs.get("wrap", False))
    if kwargs.get("border"):
        cell.border = _KENAR_TUMU
    if kwargs.get("number_format"):
        cell.number_format = kwargs["number_format"]
    return cell


# ---------------------------------------------------------------------------
# Ana fonksiyon
# ---------------------------------------------------------------------------


def hakedis_excel_olustur(
    santiye_adi: str,
    donem: str,            # "2026-05"
    is_kalemleri: list[dict],
    # Her dict: {"tanim": str, "miktar": float, "birim": str,
    #             "birim_fiyat": float, "notlar": str (opsiyonel)}
    cikti_dizin: str = "./outputs",
    kdv_orani: float = 0.20,
) -> str:
    """Hakediş Excel dosyası oluşturur ve dosya yolunu döndürür.

    Args:
        santiye_adi:  Şantiye / proje adı.
        donem:        Ay-yıl dizesi, ör. "2026-05".
        is_kalemleri: İş kalemi listesi (bkz. dict şeması yukarıda).
        cikti_dizin:  Çıktı klasörü yolu.
        kdv_orani:    KDV oranı (varsayılan 0.20 → %20).

    Returns:
        Oluşturulan Excel dosyasının tam yolu.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hakediş"
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # Satır 1 — Ana başlık
    # ------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    cell_baslik = ws["A1"]
    cell_baslik.value = f"HAKEDİŞ — {santiye_adi.upper()}"
    cell_baslik.font = Font(bold=True, size=14, color=_RENK_BEYAZ, name="Arial")
    cell_baslik.fill = _doldur(_RENK_BASLIK)
    cell_baslik.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Satır 2 — Dönem
    ws.merge_cells("A2:G2")
    cell_donem = ws["A2"]
    cell_donem.value = f"Dönem: {donem}"
    cell_donem.font = Font(bold=True, size=11, color=_RENK_BASLIK, name="Arial")
    cell_donem.fill = _doldur("E2E8F0")
    cell_donem.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Satır 3 — Boşluk
    ws.row_dimensions[3].height = 8

    # ------------------------------------------------------------------
    # Satır 4 — Tablo başlıkları
    # ------------------------------------------------------------------
    basliklar = ["#", "İş Kalemi Tanımı", "Miktar", "Birim",
                 "Birim Fiyat (₺)", "Tutar (₺)", "Notlar"]
    for col, baslik in enumerate(basliklar, 1):
        _yaz(ws, 4, col, baslik,
             bold=True, font_color=_RENK_BEYAZ, font_size=10,
             fill=_doldur(_RENK_TABLO_BAŞ), align="center", border=True)
    ws.row_dimensions[4].height = 22

    # ------------------------------------------------------------------
    # Satır 5+ — İş kalemleri
    # ------------------------------------------------------------------
    for i, kalem in enumerate(is_kalemleri, 1):
        row = 4 + i
        miktar     = kalem.get("miktar", 0) or 0
        birim_fiyat = kalem.get("birim_fiyat", 0) or 0
        tutar_ref  = f"=C{row}*E{row}"   # Excel formülü — dinamik kalır
        zebra_fill = _doldur(_RENK_ZEBRA) if i % 2 == 0 else None

        cols = [
            (1, i,                          "center"),
            (2, kalem.get("tanim", ""),     "left"),
            (3, miktar,                     "center"),
            (4, kalem.get("birim", ""),     "center"),
            (5, birim_fiyat,                "right"),
            (6, tutar_ref,                  "right"),
            (7, kalem.get("notlar", ""),    "left"),
        ]
        for col, deger, align in cols:
            kw = dict(align=align, border=True, font_size=10)
            if zebra_fill:
                kw["fill"] = zebra_fill
            cell = _yaz(ws, row, col, deger, **kw)
            if col in (5, 6):
                cell.number_format = '#,##0.00 ₺'
        ws.row_dimensions[row].height = 18

    # ------------------------------------------------------------------
    # Ara toplam, KDV, Genel Toplam
    # ------------------------------------------------------------------
    n = len(is_kalemleri)
    ara_row   = 5 + n
    kdv_row   = 6 + n
    gtop_row  = 7 + n

    def _alt_satirlar_yaz(row, etiket, formul, vurgu=False):
        ws.merge_cells(f"A{row}:E{row}")
        lbl = ws[f"A{row}"]
        lbl.value = etiket
        lbl.font = Font(bold=True, size=10, name="Arial")
        lbl.alignment = Alignment(horizontal="right", vertical="center")
        lbl.fill = _doldur(_RENK_TOPLAM_BG if not vurgu else "FDE68A")
        lbl.border = _KENAR_TUMU

        tutar = ws[f"F{row}"]
        tutar.value = formul
        tutar.font = Font(bold=True, size=10,
                          color=(_RENK_AMBER if vurgu else "000000"), name="Arial")
        tutar.fill = _doldur(_RENK_TOPLAM_BG if not vurgu else "FDE68A")
        tutar.number_format = '#,##0.00 ₺'
        tutar.alignment = Alignment(horizontal="right", vertical="center")
        tutar.border = _KENAR_TUMU
        ws.row_dimensions[row].height = 20

    if n > 0:
        ara_toplam_formul = f"=SUM(F5:F{4 + n})"
    else:
        ara_toplam_formul = 0

    _alt_satirlar_yaz(ara_row, "ARA TOPLAM", ara_toplam_formul)
    _alt_satirlar_yaz(kdv_row, f"KDV (%{int(kdv_orani * 100)})", f"=F{ara_row}*{kdv_orani}")
    _alt_satirlar_yaz(gtop_row, "GENEL TOPLAM (KDV Dahil)",
                      f"=F{ara_row}+F{kdv_row}", vurgu=True)

    # Notlar sütununu G için de kapat
    for row in (ara_row, kdv_row, gtop_row):
        ws[f"G{row}"].border = _KENAR_TUMU
        ws[f"G{row}"].fill = _doldur(_RENK_TOPLAM_BG)

    # ------------------------------------------------------------------
    # Kolon genişlikleri
    # ------------------------------------------------------------------
    genislikler = {
        "A": 5,
        "B": 42,
        "C": 10,
        "D": 10,
        "E": 16,
        "F": 16,
        "G": 22,
    }
    for harf, gen in genislikler.items():
        ws.column_dimensions[harf].width = gen

    # ------------------------------------------------------------------
    # Alt bilgi satırı
    # ------------------------------------------------------------------
    footer_row = gtop_row + 2
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = f"Bu hakediş tablosu Şantiye Asistanı tarafından otomatik oluşturulmuştur.  |  Dönem: {donem}"
    footer.font = Font(italic=True, size=9, color="94A3B8", name="Arial")
    footer.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------------------
    # Kaydet
    # ------------------------------------------------------------------
    os.makedirs(cikti_dizin, exist_ok=True)
    dosya_adi = f"hakedis_{santiye_adi.replace(' ', '_')}_{donem}.xlsx"
    dosya_yolu = os.path.join(cikti_dizin, dosya_adi)
    wb.save(dosya_yolu)
    return dosya_yolu
