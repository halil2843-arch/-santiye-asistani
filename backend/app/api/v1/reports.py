"""Rapor yönetimi endpoint'leri.

GET   /api/v1/reports/{santiye_id}?tarih=YYYY-MM-DD — günlük rapor listesi
POST  /api/v1/reports/generate                       — rapor oluştur
GET   /api/v1/reports/{rapor_id}/download            — raporu indir
PATCH /api/v1/reports/{rapor_id}/approve             — raporu onayla
"""

import asyncio
import os
from datetime import date
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import CurrentMusteriId
from app.core.database import get_db
from app.models.koordinator import Koordinator
from app.models.message import WhatsappMesaji
from app.models.report import Rapor
from app.models.site import Santiye
from app.services import rapor_servisi
from app.services.whatsapp_sender import whatsapp_bildirim_gonder

router = APIRouter(prefix="/reports", tags=["reports"])

DbDep = Annotated[AsyncSession, Depends(get_db)]


# ---------------------------------------------------------------------------
# Pydantic modelleri
# ---------------------------------------------------------------------------


class RaporResponse(BaseModel):
    """Rapor kaydının API temsili."""

    model_config = ConfigDict(from_attributes=True)

    id: str
    santiye_id: str
    sablon_id: str | None
    olusturan_id: str | None
    tarih: date
    durum: str
    cikti_dosya_yolu: str | None


class RaporOlusturRequest(BaseModel):
    """Rapor oluşturma isteği."""

    santiye_id: str
    sablon_id: str
    tarih: date
    mesaj_ids: list[str]
    proje_id: str | None = None


class RaporOlusturResponse(BaseModel):
    """Rapor oluşturma yanıtı."""

    rapor_id: str
    durum: str
    mesaj: str


# ---------------------------------------------------------------------------
# Endpoint'ler
# ---------------------------------------------------------------------------


@router.get(
    "/{santiye_id}",
    response_model=list[RaporResponse],
    summary="Günlük rapor listesi",
)
async def list_reports(
    santiye_id: str,
    db: DbDep,
    musteri_id: CurrentMusteriId,
    tarih: Annotated[date | None, Query(description="Filtre tarihi: YYYY-MM-DD")] = None,
    proje_id: Annotated[str | None, Query(description="Proje ID ile filtrele")] = None,
) -> list[RaporResponse]:
    """Şantiyenin raporlarını listeler; tarih ve/veya proje_id verilirse filtreler.

    Tenant isolation: şantiye, token sahibinin tenant'ına ait olmalıdır.

    Raises:
        404: Şantiye bulunamazsa veya başka bir tenant'a aitse.
    """
    # Şantiyenin bu tenant'a ait olduğunu doğrula
    santiye_stmt = select(Santiye).where(
        Santiye.id == santiye_id,
        Santiye.musteri_id == musteri_id,
    )
    santiye = (await db.execute(santiye_stmt)).scalar_one_or_none()
    if santiye is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Şantiye bulunamadı: {santiye_id}",
        )

    stmt = select(Rapor).where(Rapor.santiye_id == santiye_id)
    if tarih is not None:
        stmt = stmt.where(Rapor.tarih == tarih)
    if proje_id is not None and hasattr(Rapor, "proje_id"):
        stmt = stmt.where(Rapor.proje_id == proje_id)  # type: ignore[attr-defined]
    stmt = stmt.order_by(Rapor.tarih.desc())

    result = await db.execute(stmt)
    raporlar = result.scalars().all()
    return [RaporResponse.model_validate(r) for r in raporlar]


@router.post(
    "/generate",
    response_model=RaporOlusturResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Rapor oluştur (taslak)",
)
async def generate_report(
    payload: RaporOlusturRequest,
    db: DbDep,
    musteri_id: CurrentMusteriId,
) -> RaporOlusturResponse:
    """Verilen mesajlardan rapor taslağı oluşturur.

    Tenant isolation: santiye_id, token sahibinin tenant'ına ait olmalıdır.

    Raises:
        404: Şantiye bulunamazsa veya başka bir tenant'a aitse.
        404: Belirtilen mesaj ID'lerinden hiçbiri bulunamazsa.
    """
    # Şantiyenin bu tenant'a ait olduğunu doğrula
    santiye_stmt = select(Santiye).where(
        Santiye.id == payload.santiye_id,
        Santiye.musteri_id == musteri_id,
    )
    santiye = (await db.execute(santiye_stmt)).scalar_one_or_none()
    if santiye is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Şantiye bulunamadı: {payload.santiye_id}",
        )

    # Mesajları doğrula — yalnızca bu şantiyeye ait mesajlar kabul edilir
    if payload.mesaj_ids:
        stmt = select(WhatsappMesaji).where(
            WhatsappMesaji.id.in_(payload.mesaj_ids),
            WhatsappMesaji.santiye_id == payload.santiye_id,
        )
        result = await db.execute(stmt)
        mesajlar = result.scalars().all()
        if not mesajlar:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Belirtilen mesajlar bulunamadı veya bu şantiyeye ait değil.",
            )
    else:
        mesajlar = []

    rapor = Rapor(
        santiye_id=payload.santiye_id,
        sablon_id=payload.sablon_id,
        tarih=payload.tarih,
        durum="taslak",
        cikti_dosya_yolu=None,  # gerçek üretimden sonra doldurulacak
    )
    db.add(rapor)
    await db.flush()  # ID üretilsin

    # Mesajları bu raporla ilişkilendir
    for mesaj in mesajlar:
        mesaj.rapor_id = rapor.id  # type: ignore[assignment]

    await db.commit()
    await db.refresh(rapor)

    # Gerçek rapor üretimi: Groq extraction + şablon doldurma
    try:
        cikti_yolu = await rapor_servisi.uret_rapor(
            rapor_id=rapor.id, db=db, proje_id=payload.proje_id
        )
        rapor.cikti_dosya_yolu = cikti_yolu  # type: ignore[assignment]
        await db.commit()
        durum_mesaji = f"Rapor başarıyla üretildi: {cikti_yolu}"
    except asyncio.CancelledError:
        # İstemci bağlantısı kesildi — rapor taslak kalır
        durum_mesaji = "Rapor üretimi devam ediyor (istemci bağlantısı kesildi)"
        raise
    except Exception as exc:  # noqa: BLE001
        import logging
        logging.getLogger(__name__).exception("Rapor üretim hatası rapor_id=%s", rapor.id)
        rapor.durum = "hata"  # type: ignore[assignment]
        await db.commit()
        durum_mesaji = f"Üretim hatası: {exc}"

    return RaporOlusturResponse(
        rapor_id=rapor.id,
        durum=rapor.durum,
        mesaj=durum_mesaji,
    )


@router.get(
    "/{rapor_id}/download",
    summary="Rapor dosyasını indir",
)
async def download_report(
    rapor_id: str,
    db: DbDep,
    musteri_id: CurrentMusteriId,
) -> FileResponse:
    """Oluşturulmuş rapor dosyasını döndürür.

    Tenant isolation: rapor, token sahibinin tenant'ına ait şantiyeye bağlı olmalıdır.

    Raises:
        404: Rapor, çıktı dosyası bulunamazsa veya başka bir tenant'a aitse.
    """
    stmt = (
        select(Rapor)
        .join(Santiye, Rapor.santiye_id == Santiye.id)
        .where(Rapor.id == rapor_id, Santiye.musteri_id == musteri_id)
    )
    result = await db.execute(stmt)
    rapor = result.scalar_one_or_none()

    if rapor is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rapor bulunamadı: {rapor_id}",
        )

    if not rapor.cikti_dosya_yolu:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Raporun çıktı dosyası henüz oluşturulmamış.",
        )

    if not os.path.exists(rapor.cikti_dosya_yolu):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Çıktı dosyası diskte bulunamadı.",
        )

    filename = os.path.basename(rapor.cikti_dosya_yolu)
    return FileResponse(
        path=rapor.cikti_dosya_yolu,
        filename=filename,
        media_type="application/octet-stream",
    )


@router.patch(
    "/{rapor_id}/approve",
    response_model=RaporResponse,
    summary="Raporu onayla",
)
async def approve_report(
    rapor_id: str,
    db: DbDep,
    musteri_id: CurrentMusteriId,
) -> RaporResponse:
    """Rapor durumunu 'onaylandi' olarak günceller.

    Tenant isolation: rapor, token sahibinin tenant'ına ait şantiyeye bağlı olmalıdır.

    Raises:
        404: Rapor bulunamazsa veya başka bir tenant'a aitse.
        409: Rapor zaten onaylandıysa.
    """
    stmt = (
        select(Rapor)
        .join(Santiye, Rapor.santiye_id == Santiye.id)
        .where(Rapor.id == rapor_id, Santiye.musteri_id == musteri_id)
    )
    result = await db.execute(stmt)
    rapor = result.scalar_one_or_none()

    if rapor is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rapor bulunamadı: {rapor_id}",
        )

    if rapor.durum == "onaylandi":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Rapor zaten onaylanmış.",
        )

    rapor.durum = "onaylandi"  # type: ignore[assignment]
    await db.commit()
    await db.refresh(rapor)

    santiye_stmt = select(Santiye).where(Santiye.id == rapor.santiye_id)
    santiye = (await db.execute(santiye_stmt)).scalar_one_or_none()

    if santiye:
        tarih_str = rapor.tarih.strftime("%d.%m.%Y")

        if santiye.whatsapp_numara:
            asyncio.create_task(whatsapp_bildirim_gonder(
                santiye.whatsapp_numara,
                f"✅ *{santiye.isim}* — {tarih_str} tarihli rapor onaylandı. İyi çalışmalar!",
            ))

        koord_stmt = select(Koordinator).where(
            Koordinator.musteri_id == musteri_id,
            Koordinator.aktif.is_(True),
        )
        koordinatorler = (await db.execute(koord_stmt)).scalars().all()
        for k in koordinatorler:
            asyncio.create_task(whatsapp_bildirim_gonder(
                k.whatsapp_numara,
                f"✅ *{santiye.isim}* şantiyesinin {tarih_str} tarihli raporu onaylandı.",
            ))

    return RaporResponse.model_validate(rapor)


# ---------------------------------------------------------------------------
# Preview — Excel içeriğini JSON olarak döndür
# ---------------------------------------------------------------------------

class HucreDegeri(BaseModel):
    deger: str | None
    kalin: bool = False
    hizalama: str = "left"


class SayfaVeri(BaseModel):
    isim: str
    satirlar: list[list[HucreDegeri]]
    sutun_genislikleri: list[int]


class PreviewResponse(BaseModel):
    rapor_id: str
    sayfalar: list[SayfaVeri]


def _excel_oku(dosya_yolu: str) -> list[SayfaVeri]:
    import openpyxl
    wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
    sayfalar: list[SayfaVeri] = []
    for ws in wb.worksheets:
        satirlar: list[list[HucreDegeri]] = []
        sutun_genislikleri: list[int] = []
        max_satir = min(ws.max_row or 0, 100)
        max_sutun = min(ws.max_column or 0, 30)

        for r in ws.iter_rows(max_row=max_satir, max_col=max_sutun):
            satir: list[HucreDegeri] = []
            for cell in r:
                val = cell.value
                deger = str(val) if val is not None else None
                try:
                    kalin = bool(cell.font and cell.font.bold)
                except Exception:
                    kalin = False
                try:
                    hizalama = str(cell.alignment.horizontal or "left") if cell.alignment else "left"
                except Exception:
                    hizalama = "left"
                satir.append(HucreDegeri(deger=deger, kalin=kalin, hizalama=hizalama))
            satirlar.append(satir)

        sutun_genislikleri = [80] * max_sutun
        if ws.column_dimensions:
            for i, (_, cd) in enumerate(ws.column_dimensions.items()):
                if i >= max_sutun:
                    break
                try:
                    sutun_genislikleri[i] = max(40, min(300, int((cd.width or 8) * 7)))
                except Exception:
                    pass

        sayfalar.append(SayfaVeri(isim=ws.title, satirlar=satirlar, sutun_genislikleri=sutun_genislikleri))
    wb.close()
    return sayfalar


@router.get("/{rapor_id}/preview", response_model=PreviewResponse, summary="Excel raporu önizle")
async def preview_report(rapor_id: str, db: DbDep, musteri_id: CurrentMusteriId) -> PreviewResponse:
    stmt = (
        select(Rapor)
        .join(Santiye, Rapor.santiye_id == Santiye.id)
        .where(Rapor.id == rapor_id, Santiye.musteri_id == musteri_id)
    )
    rapor = (await db.execute(stmt)).scalar_one_or_none()
    if rapor is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rapor bulunamadı.")
    if not rapor.cikti_dosya_yolu or not os.path.exists(rapor.cikti_dosya_yolu):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Çıktı dosyası henüz oluşturulmamış.")
    if not rapor.cikti_dosya_yolu.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Önizleme sadece Excel (.xlsx) raporlar için destekleniyor.")

    sayfalar = await asyncio.to_thread(_excel_oku, rapor.cikti_dosya_yolu)
    return PreviewResponse(rapor_id=rapor_id, sayfalar=sayfalar)


# ---------------------------------------------------------------------------
# HTML Export — tarayıcıdan yazdırılabilir / PDF'e dönüştürülebilir
# ---------------------------------------------------------------------------


@router.get(
    "/{rapor_id}/html",
    response_class=HTMLResponse,
    summary="Raporu HTML olarak indir (tarayıcıdan yazdırılabilir)",
)
async def html_report(
    rapor_id: str,
    db: DbDep,
    musteri_id: CurrentMusteriId,
) -> HTMLResponse:
    """Oluşturulmuş Excel raporunu HTML formatında döndürür.

    Dönen sayfa tarayıcıda Ctrl+P / ⌘P ile veya 'Yazdır' butonuyla
    PDF olarak kaydedilebilir.

    Tenant isolation: rapor, token sahibinin tenant'ına ait şantiyeye bağlı olmalıdır.

    Raises:
        400: Rapor xlsx formatında değilse.
        404: Rapor veya çıktı dosyası bulunamazsa.
    """
    stmt = (
        select(Rapor)
        .join(Santiye, Rapor.santiye_id == Santiye.id)
        .where(Rapor.id == rapor_id, Santiye.musteri_id == musteri_id)
    )
    rapor = (await db.execute(stmt)).scalar_one_or_none()
    if rapor is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rapor bulunamadı: {rapor_id}",
        )
    if not rapor.cikti_dosya_yolu:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Raporun çıktı dosyası henüz oluşturulmamış.",
        )
    if not os.path.exists(rapor.cikti_dosya_yolu):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Çıktı dosyası diskte bulunamadı.",
        )
    if not rapor.cikti_dosya_yolu.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="HTML dışa aktarma sadece Excel (.xlsx) raporlar için destekleniyor.",
        )

    from app.services.pdf_exporter import excel_to_html  # noqa: PLC0415

    html_icerigi = await asyncio.to_thread(excel_to_html, rapor.cikti_dosya_yolu)
    return HTMLResponse(content=html_icerigi, status_code=200)
